from openpyxl import Workbook
from datetime import datetime
import os
def Excel(data: list, output: list, keyword: str):
    # Création d'un nouveau Workbook
    wb = Workbook()
    
    # Ajout d'une feuille à l'ouvrage
    ws = wb.active
    
    # Écriture des en-têtes dans la première ligne
    for i, header in enumerate(output, start=1):
        ws.cell(row=1, column=i, value=header)
    
    # Écriture des données dans les lignes suivantes
    for i, row in enumerate(data, start=2):
        for j, cell in enumerate(output, start=1):
            value = row.get(cell, "")
            # Convert list to string if value is a list
            if isinstance(value, list):
                value = ', '.join(map(str, value))
            ws.cell(row=i, column=j, value=value)
    
    directory_path = "../recherches/" + keyword
    
    if not os.path.exists(directory_path):
        os.makedirs(directory_path)
    
    # Enregistrement du fichier avec un nom basé sur la date actuelle
    filename = f"{directory_path}/instagram-{keyword}-{datetime.now().strftime('%Y-%m-%d')}.xlsx"
    wb.save(filename)