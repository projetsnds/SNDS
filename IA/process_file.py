import openpyxl
from openpyxl.worksheet.protection import SheetProtection
from openpyxl import load_workbook
from modele import get_modele, make_prediction
from sentiment import predict_sentiment
from nltk.sentiment import SentimentIntensityAnalyzer
from deep_translator import GoogleTranslator
import langid
from animal import analyze_animals

def categorize_data_from_file(fichier, col_entree, lst_categories, prompt=None):
    '''
    Fonction qui analyse les phrases d'un fichier .xlsx avec le réseau de neurones.
    Paramètres :
        - fichier : nom du fichier à cook (string)
        - col_entree : nom de la colonne à analyser sur chaque ligne (string)
        - lst_categories : liste des catégories de prédictions à effectuer (liste de string)
    '''
    if "environnement" in lst_categories:
        modele, tokenizer = get_modele()
    if "sentiment" in lst_categories:
        sia = SentimentIntensityAnalyzer()
        
    # indice de la colonne de chaque catégorie demandée
    columns_index = {}
    
    fichier_excel = openpyxl.load_workbook(fichier)
    feuille = fichier_excel.active
    feuille.protection.sheet = False

    for colonne in feuille.iter_cols():
        if colonne[0].value == col_entree:
            entree = colonne

    # Trouver l'index de la colonne de chaque catégorie demandée
    for i in range(len(lst_categories)):
        for colonne in feuille.iter_cols():
            if colonne[0].value == lst_categories[i]:
                columns_index[lst_categories[i]] = colonne[0].column

    # Si la colonne demandée n'existe pas, la créer
    for i in range(len(lst_categories)):
        if not lst_categories[i] in columns_index:
            sortie_index = feuille.max_column + 1
            feuille.insert_cols(sortie_index)
            feuille.cell(row=1, column=sortie_index).value = lst_categories[i]
            columns_index[lst_categories[i]] = sortie_index
            
    lst_phrases = []
            
    # Parcourir chaque ligne du fichier
    for ligne in range(2, feuille.max_row + 1):
        contenu = feuille.cell(row=ligne, column=entree[0].column).value
        
        if contenu:
            # si la phrase n'est pas en anglais :
            if langid.classify(contenu)[0] != "en":
                translation = GoogleTranslator(source='auto', target='en').translate(contenu)
            else:
                translation = contenu
        
            if "environnement" in lst_categories:
                resultat = make_prediction(modele, tokenizer, translation)
                feuille.cell(row=ligne, column=columns_index["environnement"]).value = resultat
                
            if "sentiment" in lst_categories:
                resultat = predict_sentiment(translation, sia)
                feuille.cell(row=ligne, column=columns_index["sentiment"]).value = resultat
                
            if "animal" in lst_categories and not feuille.cell(row=ligne, column=columns_index["animal"]).value:
                lst_phrases.append({"indice": ligne, "texte": translation.strip()[:50]})

    if "animal" in lst_categories:
        lst_animaux = analyze_animals(feuille, columns_index, lst_phrases, prompt)

    # Enregistrer les modifications
    fichier_excel.save(fichier)  
    fichier_excel.close()
    
    if "animal" in lst_categories:
        return lst_animaux
    else:
        return 'File processed successfully'